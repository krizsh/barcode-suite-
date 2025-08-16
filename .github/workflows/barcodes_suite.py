#!/usr/bin/env python3
# Barcodes Suite (Offline)
# Feature 1: Masterlist checking (SQLite-backed)
# Feature 2: Existing barcodes checker (two-file compare)
#
# GUI: Tkinter. Excel I/O: openpyxl. Storage: SQLite (local).
# Matches barcodes by digit runs and treats values that differ only
# by leading zeros as equal (e.g., 0123 == 123).

import os
import re
import sys
import sqlite3
from typing import Iterable, List, Set, Tuple, Dict

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
except Exception as e:
    print("Tkinter is required to run this app:", e)
    sys.exit(1)

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except Exception as e:
    print("openpyxl is required:", e)
    sys.exit(1)


APP_TITLE = "Barcodes Suite (Offline)"
DB_NAME = "barcodes_suite.db"
MASTER_TABLE = "masterlist"

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
DIGIT_RE = re.compile(r"\d{8,18}")  # 8–18 digit runs (EAN/UPC/GTIN ranges)


def normalize_barcode(code: str) -> str:
    """Normalize by stripping leading zeros so 0123 == 123. Keep '0' if all zeros."""
    s = code.lstrip("0")
    return s if s != "" else "0"


def app_dir() -> str:
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def db_path() -> str:
    return os.path.join(app_dir(), DB_NAME)


def ensure_db():
    con = sqlite3.connect(db_path())
    cur = con.cursor()
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {MASTER_TABLE} (
            barcode TEXT PRIMARY KEY
        )
    """)
    con.commit()
    con.close()


def to_str(val) -> str:
    """Convert cell value to string safely (avoid scientific notation)."""
    if val is None:
        return ""
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        as_int = int(val)
        if float(as_int) == val:
            return str(as_int)
        return "".join(ch for ch in f"{val}" if ch.isdigit())
    if isinstance(val, bytes):
        try:
            return val.decode("utf-8", "ignore")
        except:
            return str(val)
    return str(val)


def extract_barcodes_from_text(text: str) -> List[str]:
    return DIGIT_RE.findall(text or "")


def extract_barcodes_from_workbook(xlsx_path: str) -> Tuple[Set[str], Dict[str, int]]:
    wb = load_workbook(xlsx_path, data_only=True)
    found: Set[str] = set()
    counts: Dict[str, int] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                s = to_str(cell.value)
                if not s:
                    continue
                for code in extract_barcodes_from_text(s):
                    found.add(code)
                    counts[code] = counts.get(code, 0) + 1
    return found, counts


def load_masterlist_from_excel(xlsx_path: str) -> Tuple[int, int]:
    """Replace the masterlist with barcodes from Excel (normalized)."""
    ensure_db()
    found, _ = extract_barcodes_from_workbook(xlsx_path)

    con = sqlite3.connect(db_path())
    cur = con.cursor()
    cur.execute(f"DELETE FROM {MASTER_TABLE}")
    con.commit()

    inserted = 0
    skipped = 0
    for code in sorted(found):
        norm = normalize_barcode(code)
        try:
            cur.execute(f"INSERT OR IGNORE INTO {MASTER_TABLE}(barcode) VALUES (?)", (norm,))
            if cur.rowcount == 1:
                inserted += 1
            else:
                skipped += 1
        except sqlite3.Error:
            skipped += 1

    con.commit()
    con.close()
    return inserted, skipped


def get_masterlist_size() -> int:
    ensure_db()
    con = sqlite3.connect(db_path())
    cur = con.cursor()
    cur.execute(f"SELECT COUNT(*) FROM {MASTER_TABLE}")
    (n,) = cur.fetchone()
    con.close()
    return int(n or 0)


def masterlist_contains(codes: Iterable[str]) -> Set[str]:
    """Return subset of normalized codes present in masterlist."""
    ensure_db()
    norm_codes = list({normalize_barcode(c) for c in codes})
    if not norm_codes:
        return set()
    con = sqlite3.connect(db_path())
    cur = con.cursor()
    present = set()
    CHUNK = 900
    for i in range(0, len(norm_codes), CHUNK):
        chunk = norm_codes[i:i+CHUNK]
        qmarks = ",".join(["?"] * len(chunk))
        cur.execute(f"SELECT barcode FROM {MASTER_TABLE} WHERE barcode IN ({qmarks})", chunk)
        present.update(row[0] for row in cur.fetchall())
    con.close()
    return present


def highlight_matches_in_file(xlsx_path: str, out_path: str, master_present: Set[str]) -> Tuple[int, int, int]:
    wb = load_workbook(xlsx_path)
    total_occ_matches = 0
    total_occ_nonmatches = 0
    highlighted_cells = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                s = to_str(cell.value)
                if not s:
                    continue
                codes = extract_barcodes_from_text(s)
                if not codes:
                    continue
                for c in codes:
                    if normalize_barcode(c) in master_present:
                        total_occ_matches += 1
                    else:
                        total_occ_nonmatches += 1
                if any(normalize_barcode(c) in master_present for c in codes):
                    cell.fill = YELLOW_FILL
                    highlighted_cells += 1

    wb.save(out_path)
    return total_occ_matches, total_occ_nonmatches, highlighted_cells


def extract_barcodes_from_single_file(xlsx_path: str) -> Set[str]:
    found, _ = extract_barcodes_from_workbook(xlsx_path)
    return found


def remove_matches_from_second_file(existing_codes: Set[str], second_path: str, out_path: str,
                                    clear_all_highlights: bool = True) -> Tuple[int, int, int]:
    """Remove only matching barcode substrings based on normalized membership."""
    wb = load_workbook(second_path)
    occurrences_removed = 0
    occurrences_left = 0
    cells_touched = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                original_value = to_str(cell.value)
                if not original_value:
                    if clear_all_highlights and cell.fill is not None and cell.fill.fill_type is not None:
                        cell.fill = PatternFill(fill_type=None)
                    continue

                codes = extract_barcodes_from_text(original_value)
                if not codes:
                    if clear_all_highlights and cell.fill is not None and cell.fill.fill_type is not None:
                        cell.fill = PatternFill(fill_type=None)
                    continue

                new_text = original_value
                removed_here = 0
                left_here = 0
                for code in codes:
                    if normalize_barcode(code) in existing_codes:
                        import re as _re
                        new_text, nsubs = _re.subn(_re.escape(code), "", new_text)
                        removed_here += nsubs
                    else:
                        left_here += 1

                # Normalize spaces
                import re as _re2
                new_text = _re2.sub(r"\s{2,}", " ", new_text).strip()

                if removed_here > 0 or left_here > 0:
                    cells_touched += 1

                if clear_all_highlights and cell.fill is not None and cell.fill.fill_type is not None:
                    cell.fill = PatternFill(fill_type=None)

                cell.value = (None if new_text == "" else new_text)
                occurrences_removed += removed_here
                occurrences_left += left_here

    wb.save(out_path)
    return occurrences_removed, occurrences_left, cells_touched


class BarcodesSuiteApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("720x520")
        self.resizable(True, True)
        ensure_db()
        self._build_ui()
        self._refresh_master_count()

    def _build_ui(self):
        pad = 10
        nb = ttk.Notebook(self); nb.pack(fill="both", expand=True, padx=pad, pady=pad)

        # Masterlist tab
        self.tab_master = ttk.Frame(nb); nb.add(self.tab_master, text="Masterlist")
        self.lbl_master = ttk.Label(self.tab_master, text="Masterlist size: 0 barcodes", font=("Segoe UI", 10, "bold"))
        self.lbl_master.pack(anchor="w", padx=pad, pady=(pad, 4))

        btns = ttk.Frame(self.tab_master); btns.pack(fill="x", padx=pad, pady=(0, pad))
        ttk.Button(btns, text="Load/Replace Masterlist (Excel)…", command=self.on_load_master).pack(side="left", padx=(0,8))
        ttk.Button(btns, text="Clear Masterlist", command=self.on_clear_master).pack(side="left", padx=(0,8))
        ttk.Label(self.tab_master, text="Note: Only 8–18 digit runs are treated as barcodes; text is ignored.",
                  wraplength=660, foreground="#444").pack(anchor="w", padx=pad)

        # Check Against Masterlist tab
        self.tab_check = ttk.Frame(nb); nb.add(self.tab_check, text="Check Against Masterlist")
        ttk.Label(self.tab_check, text="1) Pick the Excel file to check against the stored Masterlist.",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=pad, pady=(pad,4))
        frm1 = ttk.Frame(self.tab_check); frm1.pack(fill="x", padx=pad, pady=4)
        self.entry_check_file = ttk.Entry(frm1); self.entry_check_file.pack(side="left", fill="x", expand=True)
        ttk.Button(frm1, text="Browse…", command=self.on_browse_check_file).pack(side="left", padx=(6,0))
        self.var_highlight = tk.BooleanVar(value=True)
        ttk.Checkbutton(self.tab_check, text="Highlight matches in yellow", variable=self.var_highlight).pack(anchor="w", padx=pad, pady=4)
        ttk.Button(self.tab_check, text="Run Check & Save", command=self.on_run_check).pack(anchor="w", padx=pad, pady=(6,8))
        self.txt_check_log = tk.Text(self.tab_check, height=10); self.txt_check_log.pack(fill="both", expand=True, padx=pad, pady=(0,pad))

        # Existing Barcodes Checker tab
        self.tab_twofile = ttk.Frame(nb); nb.add(self.tab_twofile, text="Existing Barcodes Checker")
        ttk.Label(self.tab_twofile, text="Compare two Excel files (Existing list vs. Second file with text+barcodes).",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=pad, pady=(pad,4))

        frma = ttk.Frame(self.tab_twofile); frma.pack(fill="x", padx=pad, pady=4)
        ttk.Label(frma, text="Existing list (Excel): ").pack(side="left")
        self.entry_existing = ttk.Entry(frma); self.entry_existing.pack(side="left", fill="x", expand=True, padx=(6,6))
        ttk.Button(frma, text="Browse…", command=self.on_browse_existing).pack(side="left")

        frmb = ttk.Frame(self.tab_twofile); frmb.pack(fill="x", padx=pad, pady=4)
        ttk.Label(frmb, text="Second file (Excel): ").pack(side="left")
        self.entry_second = ttk.Entry(frmb); self.entry_second.pack(side="left", fill="x", expand=True, padx=(6,6))
        ttk.Button(frmb, text="Browse…", command=self.on_browse_second).pack(side="left")

        self.var_clear_highlights = tk.BooleanVar(value=True)
        ttk.Checkbutton(self.tab_twofile, text="Remove highlights (cell fills) from the second file",
                        variable=self.var_clear_highlights).pack(anchor="w", padx=pad, pady=4)
        ttk.Button(self.tab_twofile, text="Run & Save Cleaned Second File", command=self.on_run_twofile).pack(anchor="w", padx=pad, pady=(6,8))
        self.txt_twofile_log = tk.Text(self.tab_twofile, height=10); self.txt_twofile_log.pack(fill="both", expand=True, padx=pad, pady=(0,pad))

        footer = ttk.Label(self, text="Offline • .xlsx only • Leading zeros are ignored for matching", foreground="#666")
        footer.pack(side="bottom", pady=(0,8))

    def _refresh_master_count(self):
        n = get_masterlist_size()
        self.lbl_master.config(text=f"Masterlist size: {n} barcodes")

    # Handlers
    def on_load_master(self):
        path = filedialog.askopenfilename(title="Choose masterlist Excel (.xlsx)", filetypes=[("Excel files", "*.xlsx")])
        if not path: return
        try:
            inserted, skipped = load_masterlist_from_excel(path)
            self._refresh_master_count()
            messagebox.showinfo("Masterlist updated", f"Inserted: {inserted}\nSkipped: {skipped}")
        except Exception as e:
            messagebox.showerror("Error loading masterlist", str(e))

    def on_clear_master(self):
        if not messagebox.askyesno("Confirm", "This will clear all barcodes from the stored masterlist. Continue?"):
            return
        try:
            ensure_db()
            con = sqlite3.connect(db_path())
            cur = con.cursor()
            cur.execute(f"DELETE FROM {MASTER_TABLE}")
            con.commit(); con.close()
            self._refresh_master_count()
            messagebox.showinfo("Cleared", "Masterlist cleared.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def on_browse_check_file(self):
        path = filedialog.askopenfilename(title="Choose Excel file to check (.xlsx)", filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.entry_check_file.delete(0, tk.END)
            self.entry_check_file.insert(0, path)

    def on_run_check(self):
        in_path = self.entry_check_file.get().strip()
        if not in_path or not os.path.exists(in_path):
            messagebox.showwarning("Missing file", "Please choose an Excel file to check.")
            return
        try:
            barcodes_in_file, occ_counts = extract_barcodes_from_workbook(in_path)
            present = masterlist_contains(barcodes_in_file)
            not_present = set(barcodes_in_file) - present

            stem, ext = os.path.splitext(in_path)
            out_path = stem + "__checked.xlsx"

            occ_matches, occ_nonmatches, cells_highlighted = (0,0,0)
            if self.var_highlight.get():
                occ_matches, occ_nonmatches, cells_highlighted = highlight_matches_in_file(in_path, out_path, present)
            else:
                wb = load_workbook(in_path); wb.save(out_path)

            log = []
            log.append(f"File: {os.path.basename(in_path)}")
            log.append(f"Unique barcodes found: {len(barcodes_in_file):,}")
            log.append(f"— In masterlist: {len(present):,}")
            log.append(f"— Not in masterlist: {len(not_present):,}")
            log.append("")
            log.append("Occurrences (per-cell barcode hits):")
            log.append(f"— Matches: {occ_matches:,}")
            log.append(f"— Non-matches: {occ_nonmatches:,}")
            log.append(f"Cells highlighted: {cells_highlighted:,}")
            log.append("")
            log.append(f"Saved: {out_path}")
            self.txt_check_log.delete("1.0", tk.END)
            self.txt_check_log.insert("1.0", "\n".join(log))
            messagebox.showinfo("Done", f"Saved output:\n{out_path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def on_browse_existing(self):
        path = filedialog.askopenfilename(title="Choose existing barcode list (.xlsx)", filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.entry_existing.delete(0, tk.END)
            self.entry_existing.insert(0, path)

    def on_browse_second(self):
        path = filedialog.askopenfilename(title="Choose second file (.xlsx) with text and barcodes", filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.entry_second.delete(0, tk.END)
            self.entry_second.insert(0, path)

    def on_run_twofile(self):
        first = self.entry_existing.get().strip()
        second = self.entry_second.get().strip()
        if not first or not os.path.exists(first):
            messagebox.showwarning("Missing file", "Please choose the existing barcode list (.xlsx).")
            return
        if not second or not os.path.exists(second):
            messagebox.showwarning("Missing file", "Please choose the second file (.xlsx).")
            return
        try:
            existing_codes = {normalize_barcode(c) for c in extract_barcodes_from_single_file(first)}

            stem, ext = os.path.splitext(second)
            out_path = stem + "__cleaned.xlsx"

            occ_removed, occ_left, cells_touched = remove_matches_from_second_file(
                existing_codes, second, out_path, clear_all_highlights=self.var_clear_highlights.get()
            )

            log = []
            log.append(f"Existing list: {os.path.basename(first)} — unique barcodes: {len(existing_codes):,}")
            log.append(f"Second file: {os.path.basename(second)}")
            log.append("")
            log.append(f"Occurrences removed (matched existing): {occ_removed:,}")
            log.append(f"Occurrences kept (not in existing): {occ_left:,}")
            log.append(f"Cells touched: {cells_touched:,}")
            log.append("")
            log.append(f"Saved: {out_path}")
            self.txt_twofile_log.delete("1.0", tk.END)
            self.txt_twofile_log.insert("1.0", "\n".join(log))
            messagebox.showinfo("Done", f"Saved output:\n{out_path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))


def main():
    root = BarcodesSuiteApp()
    root.mainloop()


if __name__ == "__main__":
    main()
