#!/usr/bin/env python3
# Barcodes Suite (Offline) — v8
# - Reads barcodes separated by comma/semicolon/dash as individual codes.
# - Joins hyphens inside a single code (no spaces) so "978-0-..." becomes one barcode.
# - Tidier cleanup after removals.
# - Includes v7 features (reset buttons, per-column summaries, strong clearing, toggle).

import os, re, sys, csv, sqlite3
from typing import Iterable, List, Set, Tuple, Dict

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
except Exception as e:
    print("Tkinter is required to run this app:", e)
    sys.exit(1)

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill
except Exception as e:
    print("openpyxl is required:", e)
    sys.exit(1)

try:
    import xlrd  # for .xls (v1.2.0)
    _XLRD_OK = True
except Exception:
    _XLRD_OK = False

APP_TITLE = "Barcodes Suite (Offline)"
DB_NAME = "barcodes_suite.db"
MASTER_TABLE = "masterlist"

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
DIGIT_RE = re.compile(r"\d{5,20}")
MIN_LEN, MAX_LEN = 5, 14

RE_HYPHEN_JOIN = re.compile(r"(?<=\d)-(?!\s)(?=\d)")
RE_DASHES = re.compile(r"[–—-]")
RE_SEP = re.compile(r"\s*[,;–—-]\s*")


def normalize_barcode(code: str) -> str:
    s = "".join(ch for ch in str(code) if ch.isdigit()).lstrip("0")
    return s if s != "" else "0"


def is_valid_len_after_norm(code: str) -> bool:
    n = len(normalize_barcode(code))
    return MIN_LEN <= n <= MAX_LEN


def app_dir() -> str:
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def db_path() -> str:
    return os.path.join(app_dir(), DB_NAME)


def ensure_db():
    con = sqlite3.connect(db_path())
    cur = con.cursor()
    cur.execute(f"CREATE TABLE IF NOT EXISTS {MASTER_TABLE} (barcode TEXT PRIMARY KEY)")
    con.commit(); con.close()


def to_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        as_int = int(val)
        if float(as_int) == val:
            return str(as_int)
        return "".join(ch for ch in f"{val}" if ch.isdigit())
    if isinstance(val, bytes):
        try: return val.decode("utf-8", "ignore")
        except: return str(val)
    return str(val)


def load_workbook_generic(path: str):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return load_workbook(path, data_only=True)
    if ext == ".csv":
        wb = Workbook()
        ws = wb.active
        ws.title = os.path.splitext(os.path.basename(path))[0][:31] or "CSV"
        with open(path, "r", encoding="utf-8", newline="") as f:
            for row in csv.reader(f):
                ws.append(row)
        return wb
    if ext == ".xls":
        if not _XLRD_OK:
            raise RuntimeError("Reading .xls requires xlrd==1.2.0")
        book = xlrd.open_workbook(path)
        wb = Workbook()
        if wb.worksheets: wb.remove(wb.worksheets[0])
        for s in book.sheets():
            ws = wb.create_sheet(title=str(s.name)[:31])
            for r in range(s.nrows):
                ws.append([s.cell_value(r, c) for c in range(s.ncols)])
        return wb
    return load_workbook(path, data_only=True)


def extract_barcodes_from_text(text: str) -> List[str]:
    if not text:
        return []
    s = RE_DASHES.sub("-", str(text))   # normalize – and — to -
    s = RE_HYPHEN_JOIN.sub("", s)       # join 978-0-... -> 9780...
    return DIGIT_RE.findall(s)


def extract_barcodes_from_workbook_any(path: str) -> Tuple[Set[str], Dict[str, int]]:
    wb = load_workbook_generic(path)
    found: Set[str] = set()
    counts: Dict[str, int] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                s = to_str(cell.value)
                if not s: continue
                for raw in extract_barcodes_from_text(s):
                    if not is_valid_len_after_norm(raw): continue
                    norm = normalize_barcode(raw)
                    found.add(norm)
                    counts[norm] = counts.get(norm, 0) + 1
    return found, counts


def load_masterlist_from_file(path: str) -> Tuple[int, int]:
    ensure_db()
    found, _ = extract_barcodes_from_workbook_any(path)
    con = sqlite3.connect(db_path()); cur = con.cursor()
    cur.execute(f"DELETE FROM {MASTER_TABLE}"); con.commit()
    ins=0; skip=0
    for norm in sorted(found):
        try:
            cur.execute(f"INSERT OR IGNORE INTO {MASTER_TABLE}(barcode) VALUES (?)", (norm,))
            ins += (1 if cur.rowcount==1 else 0); skip += (0 if cur.rowcount==1 else 1)
        except sqlite3.Error:
            skip += 1
    con.commit(); con.close()
    return ins, skip


def get_masterlist_size() -> int:
    ensure_db()
    con = sqlite3.connect(db_path()); cur = con.cursor()
    cur.execute(f"SELECT COUNT(*) FROM {MASTER_TABLE}"); (n,) = cur.fetchone()
    con.close(); return int(n or 0)


def masterlist_contains(codes: Iterable[str]) -> Set[str]:
    ensure_db()
    norm_codes = list({normalize_barcode(c) for c in codes if is_valid_len_after_norm(c)})
    if not norm_codes: return set()
    con = sqlite3.connect(db_path()); cur = con.cursor()
    present=set(); CHUNK=900
    for i in range(0, len(norm_codes), CHUNK):
        chunk = norm_codes[i:i+CHUNK]
        qmarks = ",".join(["?"]*len(chunk))
        cur.execute(f"SELECT barcode FROM {MASTER_TABLE} WHERE barcode IN ({qmarks})", chunk)
        present.update(row[0] for row in cur.fetchall())
    con.close(); return present


def clear_all_fills_and_rules(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if getattr(cell, "fill", None) is not None:
                    cell.fill = PatternFill(fill_type=None)
        try:
            cf = getattr(ws, "conditional_formatting", None)
            if cf is not None:
                if hasattr(cf, "clear"):
                    cf.clear()  # type: ignore
                elif hasattr(cf, "_cf_rules"):
                    cf._cf_rules.clear()  # type: ignore
                elif hasattr(cf, "cf_rules"):
                    cf.cf_rules.clear()  # type: ignore
        except Exception:
            pass


def used_columns_count(ws) -> int:
    used=set()
    for row in ws.iter_rows():
        for cell in row:
            v = to_str(cell.value).strip()
            if v != "":
                used.add(cell.column)
    return len(used)


def process_masterlist_check(in_path: str, out_path: str, master_present: Set[str], do_highlight: bool) -> Tuple[int,int,int]:
    wb = load_workbook_generic(in_path)
    clear_all_fills_and_rules(wb)

    total_occ_matches=0; total_occ_unmatches=0; total_cells_highlighted=0

    for ws in wb.worksheets:
        occ_m=0; occ_u=0; cells_h=0
        single_col = (used_columns_count(ws) <= 1)

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        col_match = [0]*(max_col+1)
        col_unmatch = [0]*(max_col+1)

        if single_col:
            ws.cell(row=1, column=2, value="Status")

        for r in range(1, max_row+1):
            for c in range(1, max_col+1):
                cell = ws.cell(row=r, column=c)
                s = to_str(cell.value)
                if not s: continue
                raw_codes = extract_barcodes_from_text(s)
                if not raw_codes: continue
                norm_codes = [normalize_barcode(x) for x in raw_codes if is_valid_len_after_norm(x)]
                if not norm_codes: continue

                any_match = any(nc in master_present for nc in norm_codes)
                for nc in norm_codes:
                    if nc in master_present:
                        occ_m += 1; total_occ_matches += 1; col_match[c] += 1
                    else:
                        occ_u += 1; total_occ_unmatches += 1; col_unmatch[c] += 1

                if do_highlight and any_match:
                    cell.fill = YELLOW_FILL
                    cells_h += 1; total_cells_highlighted += 1

                if single_col and c == 1:
                    ws.cell(row=r, column=2, value=("match" if any_match else "unmatch"))

        if not single_col and max_col > 0:
            ws.insert_rows(1, 2)
            for c in range(1, max_col+1):
                ws.cell(row=1, column=c, value=f"unmatch: {col_unmatch[c]}")
                ws.cell(row=2, column=c, value=f"match: {col_match[c]}")

    wb.save(out_path)
    return total_occ_matches, total_occ_unmatches, total_cells_highlighted


def extract_barcodes_from_single_file(path: str) -> Set[str]:
    found, _ = extract_barcodes_from_workbook_any(path)
    return found


def tidy_text_after_removal(text: str) -> str:
    if not text:
        return text
    s = RE_DASHES.sub("-", str(text))
    s = re.sub(r"^[\s,;–—-]+|[\s,;–—-]+$", "", s)
    s = re.sub(r"\s*[,;–—-]\s*([,;–—-]\s*)+", ", ", s)
    s = re.sub(r"(^|[\s])[,;–—-]\s*", r"\1", s)
    s = re.sub(r"\s*[-;]\s*", ", ", s)
    s = re.sub(r"\s{2,}", " ", s).strip()
    s = re.sub(r"^,\s*|\s*,\s*$", "", s)
    return s


def remove_matches_existing(existing_codes: Set[str], second_path: str, out_path: str) -> Tuple[int,int,int,int]:
    wb = load_workbook_generic(second_path)
    clear_all_fills_and_rules(wb)

    occ_removed=0; occ_left=0; cells_touched=0; rows_deleted_total=0

    for ws in wb.worksheets:
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        col_existing = [0]*(max_col+1)
        col_nonexisting = [0]*(max_col+1)

        for r in range(1, max_row+1):
            for c in range(1, max_col+1):
                cell = ws.cell(row=r, column=c)
                original_value = to_str(cell.value)
                if not original_value: continue
                raw_codes = extract_barcodes_from_text(original_value)
                if not raw_codes: continue

                new_text = original_value
                removed_here = 0; left_here = 0
                for raw in raw_codes:
                    if not is_valid_len_after_norm(raw):
                        left_here += 1; continue
                    norm = normalize_barcode(raw)
                    if norm in existing_codes:
                        import re as _re
                        new_text, nsubs = _re.subn(_re.escape(raw), "", new_text)
                        removed_here += nsubs
                    else:
                        left_here += 1

                new_text = tidy_text_after_removal(new_text)

                if removed_here > 0 or left_here > 0:
                    cells_touched += 1

                if getattr(cell, "fill", None) is not None:
                    cell.fill = PatternFill(fill_type=None)

                cell.value = (None if new_text == "" else new_text)
                occ_removed += removed_here; occ_left += left_here
                col_existing[c] += removed_here
                col_nonexisting[c] += left_here

        empties = []
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        for r in range(1, max_row+1):
            all_empty = True
            for c in range(1, max_col+1):
                v = to_str(ws.cell(row=r, column=c).value).strip()
                if v != "":
                    all_empty = False; break
            if all_empty: empties.append(r)
        for r in reversed(empties):
            ws.delete_rows(r, 1); rows_deleted_total += 1

        if max_col > 1:
            ws.insert_rows(1, 2)
            for c in range(1, max_col+1):
                ws.cell(row=1, column=c, value=f"non-existing: {col_nonexisting[c]}")
                ws.cell(row=2, column=c, value=f"existing: {col_existing[c]}")

    wb.save(out_path)
    return occ_removed, occ_left, cells_touched, rows_deleted_total


class BarcodesSuiteApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("900x640")
        self.resizable(True, True)
        ensure_db()
        self._build_ui()
        self._refresh_master_count()

    def _build_ui(self):
        pad = 10
        nb = ttk.Notebook(self); nb.pack(fill="both", expand=True, padx=pad, pady=pad)

        self.tab_master = ttk.Frame(nb); nb.add(self.tab_master, text="Masterlist")
        self.lbl_master = ttk.Label(self.tab_master, text="Masterlist size: 0 barcodes", font=("Segoe UI", 10, "bold"))
        self.lbl_master.pack(anchor="w", padx=pad, pady=(pad, 4))
        btns = ttk.Frame(self.tab_master); btns.pack(fill="x", padx=pad, pady=(0, pad))
        ttk.Button(btns, text="Load/Replace Masterlist (Excel/CSV)…", command=self.on_load_master).pack(side="left", padx=(0,8))
        ttk.Button(btns, text="Clear Masterlist", command=self.on_clear_master).pack(side="left", padx=(0,8))
        ttk.Label(self.tab_master, text="Reads .xlsx/.xlsm/.xls/.csv • Only 5–14 digits treated as barcodes (leading zeros ignored).",
                  wraplength=860, foreground="#444").pack(anchor="w", padx=pad)

        self.tab_check = ttk.Frame(nb); nb.add(self.tab_check, text="Check Against Masterlist")
        ttk.Label(self.tab_check, text="Pick a file to check against the stored Masterlist.",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=pad, pady=(pad,4))
        frm1 = ttk.Frame(self.tab_check); frm1.pack(fill="x", padx=pad, pady=4)
        self.entry_check_file = ttk.Entry(frm1); self.entry_check_file.pack(side="left", fill="x", expand=True)
        ttk.Button(frm1, text="Browse…", command=self.on_browse_check_file).pack(side="left", padx=(6,0))
        self.var_highlight = tk.BooleanVar(value=True)
        ttk.Checkbutton(self.tab_check, text="Highlight matches in yellow", variable=self.var_highlight).pack(anchor="w", padx=pad, pady=4)
        act = ttk.Frame(self.tab_check); act.pack(fill="x", padx=pad, pady=(0,6))
        ttk.Button(act, text="Run Check & Save", command=self.on_run_check).pack(side="left")
        ttk.Button(act, text="Reset", command=self.on_reset_check).pack(side="left", padx=(8,0))
        self.txt_check_log = tk.Text(self.tab_check, height=12); self.txt_check_log.pack(fill="both", expand=True, padx=pad, pady=(0,pad))

        self.tab_twofile = ttk.Frame(nb); nb.add(self.tab_twofile, text="Existing Barcodes Checker")
        ttk.Label(self.tab_twofile, text="Existing list vs. Second file (text + barcodes).",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=pad, pady=(pad,4))
        frma = ttk.Frame(self.tab_twofile); frma.pack(fill="x", padx=pad, pady=4)
        ttk.Label(frma, text="Existing list (Excel/CSV): ").pack(side="left")
        self.entry_existing = ttk.Entry(frma); self.entry_existing.pack(side="left", fill="x", expand=True, padx=(6,6))
        ttk.Button(frma, text="Browse…", command=self.on_browse_existing).pack(side="left")
        frmb = ttk.Frame(self.tab_twofile); frmb.pack(fill="x", padx=pad, pady=4)
        ttk.Label(frmb, text="Second file (Excel/CSV): ").pack(side="left")
        self.entry_second = ttk.Entry(frmb); self.entry_second.pack(side="left", fill="x", expand=True, padx=(6,6))
        ttk.Button(frmb, text="Browse…", command=self.on_browse_second).pack(side="left")
        act2 = ttk.Frame(self.tab_twofile); act2.pack(fill="x", padx=pad, pady=(0,6))
        ttk.Button(act2, text="Run & Save Cleaned Second File", command=self.on_run_twofile).pack(side="left")
        ttk.Button(act2, text="Reset", command=self.on_reset_twofile).pack(side="left", padx=(8,0))
        self.txt_twofile_log = tk.Text(self.tab_twofile, height=12); self.txt_twofile_log.pack(fill="both", expand=True, padx=pad, pady=(0,pad))

        footer = ttk.Label(self, text="Offline • Reads .xlsx .xlsm .xls .csv • Leading zeros ignored • 5–14 digits", foreground="#666")
        footer.pack(side="bottom", pady=(0,8))

    def _refresh_master_count(self):
        self.lbl_master.config(text=f"Masterlist size: {get_masterlist_size()} barcodes")

    def on_load_master(self):
        path = filedialog.askopenfilename(title="Choose Masterlist file",
                                          filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.xls *.csv"), ("All files","*.*")])
        if not path: return
        try:
            ins, skip = load_masterlist_from_file(path)
            self._refresh_master_count()
            messagebox.showinfo("Masterlist updated", f"Inserted: {ins}\nSkipped: {skip}")
        except Exception as e:
            messagebox.showerror("Error loading masterlist", str(e))

    def on_clear_master(self):
        if not messagebox.askyesno("Confirm", "Clear all barcodes from the stored masterlist?"): return
        try:
            ensure_db(); con = sqlite3.connect(db_path()); cur = con.cursor()
            cur.execute(f"DELETE FROM {MASTER_TABLE}"); con.commit(); con.close()
            self._refresh_master_count(); messagebox.showinfo("Cleared", "Masterlist cleared.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def on_browse_check_file(self):
        path = filedialog.askopenfilename(title="Choose file to check",
                                          filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.xls *.csv"), ("All files","*.*")])
        if path:
            self.entry_check_file.delete(0, tk.END); self.entry_check_file.insert(0, path)

    def on_reset_check(self):
        self.entry_check_file.delete(0, tk.END)
        self.var_highlight.set(True)
        self.txt_check_log.delete("1.0", tk.END)

    def on_run_check(self):
        in_path = self.entry_check_file.get().strip()
        if not in_path or not os.path.exists(in_path):
            messagebox.showwarning("Missing file", "Please choose a file to check."); return
        try:
            barcodes_in_file, occ_counts = extract_barcodes_from_workbook_any(in_path)
            present = masterlist_contains(barcodes_in_file)
            not_present = set(barcodes_in_file) - present
            stem, _ = os.path.splitext(in_path); out_path = stem + "__checked.xlsx"
            occ_m, occ_u, cells_h = process_masterlist_check(in_path, out_path, present, self.var_highlight.get())

            log=[f"File: {os.path.basename(in_path)}",
                 f"Unique barcodes found (5–14 digits, normalized): {len(barcodes_in_file):,}",
                 f"— match: {len(present):,}",
                 f"— unmatch: {len(not_present):,}", "",
                 "Occurrences (per-cell barcode hits):",
                 f"— match: {occ_m:,}",
                 f"— unmatch: {occ_u:,}",
                 f"Cells highlighted: {cells_h:,}",
                 f"Highlight setting: {'ON' if self.var_highlight.get() else 'OFF'}", "",
                 f"Saved: {out_path}"]
            self.txt_check_log.delete("1.0", tk.END); self.txt_check_log.insert("1.0", "\n".join(log))
            messagebox.showinfo("Done", f"Saved output:\n{out_path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def on_browse_existing(self):
        path = filedialog.askopenfilename(title="Choose existing barcode list",
                                          filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.xls *.csv"), ("All files","*.*")])
        if path:
            self.entry_existing.delete(0, tk.END); self.entry_existing.insert(0, path)

    def on_browse_second(self):
        path = filedialog.askopenfilename(title="Choose second file (text + barcodes)",
                                          filetypes=[("Excel/CSV", "*.xlsx *.xlsm *.xls *.csv"), ("All files","*.*")])
        if path:
            self.entry_second.delete(0, tk.END); self.entry_second.insert(0, path)

    def on_reset_twofile(self):
        self.entry_existing.delete(0, tk.END)
        self.entry_second.delete(0, tk.END)
        self.txt_twofile_log.delete("1.0", tk.END)

    def on_run_twofile(self):
        first = self.entry_existing.get().strip(); second = self.entry_second.get().strip()
        if not first or not os.path.exists(first):
            messagebox.showwarning("Missing file", "Choose the existing barcode list."); return
        if not second or not os.path.exists(second):
            messagebox.showwarning("Missing file", "Choose the second file."); return
        try:
            existing_codes = extract_barcodes_from_single_file(first)
            stem, _ = os.path.splitext(second); out_path = stem + "__cleaned.xlsx"
            occ_r, occ_l, cells_t, rows_del = remove_matches_existing(existing_codes, second, out_path)
            log=[f"Existing list: {os.path.basename(first)} — unique barcodes: {len(existing_codes):,}",
                 f"Second file: {os.path.basename(second)}", "",
                 f"Occurrences removed (existing): {occ_r:,}",
                 f"Occurrences kept (non-existing): {occ_l:,}",
                 f"Cells touched: {cells_t:,}",
                 f"Blank rows deleted: {rows_del:,}", "",
                 f"Saved: {out_path}"]
            self.txt_twofile_log.delete("1.0", tk.END); self.txt_twofile_log.insert("1.0", "\n".join(log))
            messagebox.showinfo("Done", f"Saved output:\n{out_path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))


def main():
    root = BarcodesSuiteApp(); root.mainloop()

if __name__ == "__main__":
    main()
